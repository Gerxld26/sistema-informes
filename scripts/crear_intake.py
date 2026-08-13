from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

RAIZ = Path(__file__).resolve().parent.parent
RUTA_SALIDA = RAIZ / "intake" / "intake_informes.xlsx"

wb = openpyxl.Workbook()

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Arial", size=10)
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
NOTE_FONT = Font(italic=True, color="808080", name="Arial", size=9)


def style_header(ws, row=1, ncols=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ============ HOJA 1: TRAMITES ============
ws1 = wb.active
ws1.title = "Tramites"
headers1 = [
    "id_tramite", "numero_informe", "anio", "tipo_tramite",
    "siglas_sede", "fecha", "tecnico_nombre", "tecnico_cargo",
    "destinatario_nombre", "destinatario_cargo",
    "jefe_ti_nombre", "jefe_ti_cargo",
]
ws1.append(headers1)
style_header(ws1, ncols=len(headers1))

ws1.append([
    "T001", 106, 2026, "implementacion", "CHSP", "23/04/2026",
    "Antony Reyes Recalde", "Soporte Tecnico",
    "Grover Loayza", "Jefe Corporativo TI",
    "Stefan Mandujano", "Jefe TI Surco",
])
ws1.append([
    "T002", 107, 2026, "baja", "LMP", "10/08/2026",
    "Luis Mendoza Paredes", "Soporte Tecnico",
    "Grover Loayza", "Jefe Corporativo TI",
    "Stefan Mandujano", "Jefe TI Surco",
])

widths1 = [12, 15, 8, 15, 12, 12, 22, 18, 22, 18, 20, 16]
for i, w in enumerate(widths1, start=1):
    ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws1.append([])
nota1 = ws1.cell(row=ws1.max_row + 1, column=1, value=(
    "Nota: tipo_tramite = baja | implementacion | diagnostico. "
    "siglas_sede: sigla que va despues de las iniciales del tecnico (ej: ARR/CHSP). "
    "id_tramite debe ser unico y se referencia desde la hoja Equipos."
))
nota1.font = NOTE_FONT

dv_tipo = DataValidation(type="list", formula1='"baja,implementacion,diagnostico"', allow_blank=False)
ws1.add_data_validation(dv_tipo)
dv_tipo.add("D2:D500")

# ============ HOJA 2: EQUIPOS ============
ws2 = wb.create_sheet("Equipos")
headers2 = [
    "id_tramite", "usuario", "cargo_usuario", "area", "sede_usuario", "jefatura",
    "tipo_bien", "equipo_actual_marca_estado", "sede_equipo_actual",
    "texto_justificacion_override", "especificaciones_override",
]
ws2.append(headers2)
style_header(ws2, ncols=len(headers2))

ws2.append([
    "T001", "Edison Cardenas Masias", "Analista", "Farmacia Corporativa",
    "San Juan Bautista", "Alejandro Izquierdo", "CPU",
    "equipo desfasado en uso, con problemas de lentitud general", "Chacarilla",
    "", "",
])
ws2.append([
    "T002", "Maria Torres Vega", "Analista", "Contabilidad", "Surco",
    "Roberto Diaz", "CPU", "con 8 anios de antiguedad, ya no soporta el ERP", "Surco", "", "",
])
ws2.append([
    "T002", "Jose Ramos Chavez", "Gerente", "Contabilidad", "Surco",
    "Roberto Diaz", "Laptop", "pantalla danada y bateria agotada", "Surco", "", "",
])

widths2 = [12, 24, 14, 22, 16, 20, 10, 34, 18, 40, 40]
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws2.append([])
nota2 = ws2.cell(row=ws2.max_row + 1, column=1, value=(
    "Nota: 1 fila = 1 equipo. Para un tramite con 5 equipos, se repiten 5 filas con el mismo "
    "id_tramite. Deja 'especificaciones_override' vacio para que el sistema use el Catalogo "
    "segun area/cargo; llenalo (separado por ; ) solo para casos especiales (ej: gerentes)."
))
nota2.font = NOTE_FONT

dv_bien = DataValidation(type="list", formula1='"CPU,Laptop,Monitor,Impresora"', allow_blank=False)
ws2.add_data_validation(dv_bien)
dv_bien.add("F2:F500")

# ============ HOJA 3: CATALOGO_EQUIPOS ============
ws3 = wb.create_sheet("Catalogo_Equipos")
headers3 = [
    "prioridad", "campo", "valor", "tipo_bien", "especificaciones (separadas por ;)",
]
ws3.append(headers3)
style_header(ws3, ncols=len(headers3))

catalogo_rows = [
    [1, "cargo", "Gerente", "CPU",
     "HP Core i7 14va generacion;SSD Kingston NVME M.2 1TB;32GB RAM;Licencia Windows 11 Professional;Case;Monitor 27\""],
    [1, "cargo", "Jefatura", "CPU",
     "HP Core i7 14va generacion;SSD Kingston NVME M.2 1TB;16GB RAM;Licencia Windows 11 Professional;Case"],
    [2, "area", "Farmacia Corporativa", "CPU",
     "HP Core i5 14va generacion;SSD Kingston NVME M.2 1TB;16GB RAM;Licencia Windows 11 Professional;Case"],
    [2, "area", "Sistemas", "CPU",
     "HP Core i7 14va generacion;SSD Kingston NVME M.2 1TB;16GB RAM;Licencia Windows 11 Professional;Case"],
    [3, "default", "*", "CPU",
     "HP Core i3 14va generacion;SSD Kingston NVME M.2 512GB;8GB RAM;Licencia Windows 11 Professional;Case"],
    [3, "default", "*", "Laptop",
     "HP ProBook 14 pulg;SSD Kingston NVME M.2 512GB;8GB RAM;Licencia Windows 11 Professional"],
]
for row in catalogo_rows:
    ws3.append(row)

widths3 = [10, 12, 22, 12, 70]
for i, w in enumerate(widths3, start=1):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws3.append([])
nota3 = ws3.cell(row=ws3.max_row + 1, column=1, value=(
    "Nota: 'prioridad' mas baja = se evalua primero (mas especifico). El sistema busca primero "
    "coincidencia por cargo, luego por area, y si no encuentra nada usa la fila 'default' segun "
    "tipo_bien. Editen esta hoja libremente sin tocar el script."
))
nota3.font = NOTE_FONT

# ============ HOJA 4: RESPONSABLES ============
ws4 = wb.create_sheet("Responsables")
headers4 = ["area", "etiqueta_responsable", "nombre_responsable"]
ws4.append(headers4)
style_header(ws4, ncols=len(headers4))

responsables_rows = [
    ["Farmacia Corporativa", "JEFE", "ALEJANDRO IZQUIERDO"],
    ["Contabilidad", "A CARGO", "ROBERTO DIAZ"],
    ["Sistemas", "JEFATURA", "STEFAN MANDUJANO"],
    ["CSO", "JEFE", "GIANNINA DOMINGUEZ"],
]
for row in responsables_rows:
    ws4.append(row)

widths4 = [22, 18, 24]
for i, w in enumerate(widths4, start=1):
    ws4.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws4.append([])
nota4 = ws4.cell(row=ws4.max_row + 1, column=1, value=(
    "Nota: al elegir un area en el wizard, se sugiere automaticamente el responsable de esta "
    "tabla (editable en el momento). Si el area no esta aqui, el wizard pide los datos a mano."
))
nota4.font = NOTE_FONT

RUTA_SALIDA.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(RUTA_SALIDA))
print(f"Intake creado en {RUTA_SALIDA}")
